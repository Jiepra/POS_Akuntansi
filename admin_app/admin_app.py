import sys
import mysql.connector
import datetime
import tempfile
import os 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QTabWidget, QTableWidget, QTableWidgetItem,
                             QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QFormLayout, 
                             QDateEdit, QStatusBar, QHeaderView, QMessageBox, QFileDialog, QSizePolicy) # <-- QSizePolicy dipindahkan ke sini
from PyQt5.QtCore import Qt, QDate # QSizePolicy dihapus dari sini
from PyQt5.QtGui import QFont, QColor

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# Konfigurasi database
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'ilhamDB26',
    'database': 'akuntansi'
}

def connect_db():
    """Membuat koneksi ke database MySQL."""
    return mysql.connector.connect(**DB_CONFIG)

class AdminApp(QMainWindow):
    """
    Aplikasi desktop untuk manajemen admin, termasuk manajemen produk,
    tampilan transaksi, dan laporan keuangan.
    """
    def __init__(self):
        super().__init__()
        self.report_data = {
            'total_penjualan': 0,
            'total_pembelian': 0,
            'laba_kotor': 0,
            'laba_bersih': 0,
            'start_date': '',
            'end_date': '',
            'periode': ''
        }
        
        self.setup_ui()
        self.load_products()
        self.set_initial_dates()
        self.filter_transactions()
        self.update_report()
        
    def setup_ui(self):
        """Mengatur tampilan utama dan widget aplikasi."""
        self.setWindowTitle("Admin Dashboard - Toko Modern")
        self.setGeometry(100, 100, 1400, 900) 
        self.setStyleSheet(self.get_stylesheet())
        
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
        
        self.create_product_tab()
        self.create_transaction_tab()
        self.create_report_tab()
        
        
        self.tabs.setTabText(0, "Manajemen Produk") 
        self.tabs.setTabText(1, "Transaksi")
        self.tabs.setTabText(2, "Laporan Keuangan") 
        
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")
    
    def get_stylesheet(self):
        """Mengembalikan stylesheet CSS untuk aplikasi."""
        return """
            QMainWindow { background-color: #f0f0f0; }
            QTabWidget::pane { 
                border: 1px solid #d0d0d0; 
                background: white; 
                margin-top: 10px; 
                border-radius: 10px; 
            }
            QTabBar::tab {
                background: #e0e0e0; 
                border: 1px solid #d0d0d0;
                padding: 10px 25px; 
                font-weight: bold; 
                color: #404040;
                min-width: 200px; 
                border-top-left-radius: 8px; 
                border-top-right-radius: 8px;
            }
            QTabBar::tab:selected {
                background: white; 
                border-bottom: 2px solid #2196F3; 
                color: #2196F3;
            }
            QTableWidget { 
                background: white; 
                border: 1px solid #d0d0d0; 
                font-size: 14px; 
                gridline-color: #e0e0e0; 
                border-radius: 8px; 
            }
            QTableWidget::item { 
                padding: 14px; 
            }
            QHeaderView::section {
                background-color: #2196F3; 
                color: white; 
                font-weight: bold;
                padding: 10px; 
                border: none;
            }
            QLineEdit { 
                padding: 13px; 
                border: 1px solid #d0d0d0; 
                border-radius: 6px; 
                font-size: 18px; 
            }
            QLabel {
                font-size: 18px; 
                color: #333333;
            }
            QDateEdit {
                padding: 18px;
                border: 1px solid #d0d0d0;
                border-radius: 6px;
                font-size: 15px;
            }
        """
        
    def set_initial_dates(self):
        """Mengatur tanggal awal untuk filter transaksi dan laporan."""
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        self.date_to.setDate(QDate.currentDate())
        
        self.report_date_from.setDate(QDate.currentDate().addDays(-30))
        self.report_date_to.setDate(QDate.currentDate())
        
        self.report_data.update({
            'start_date': self.report_date_from.date().toString("yyyy-MM-dd"),
            'end_date': self.report_date_to.date().toString("yyyy-MM-dd"),
            'periode': f"{self.report_date_from.date().toString('dd/MM/yyyy')} - {self.report_date_to.date().toString('dd/MM/yyyy')}"
        })

    # --- Tab Manajemen Produk ---
    def create_product_tab(self):
        """Membuat tab untuk manajemen produk."""
        product_tab = QWidget()
        layout = QVBoxLayout(product_tab)
        layout.setContentsMargins(30, 30, 30, 30) 
        layout.setSpacing(20) 
        
        form_layout = QFormLayout()
        self.kode_input = QLineEdit()
        self.nama_input = QLineEdit()
        self.stok_input = QLineEdit()
        self.harga_beli_input = QLineEdit()
        self.harga_jual_input = QLineEdit()
        
        form_layout.setLabelAlignment(Qt.AlignRight)
        self.kode_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.nama_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.stok_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.harga_beli_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.harga_jual_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        form_layout.addRow("Kode:", self.kode_input)
        form_layout.addRow("Nama:", self.nama_input)
        form_layout.addRow("Stok:", self.stok_input)
        form_layout.addRow("Harga Beli:", self.harga_beli_input)
        form_layout.addRow("Harga Jual:", self.harga_jual_input)
        
        btn_layout = QHBoxLayout()
        self.tambah_btn = QPushButton("Tambah")
        self.update_btn = QPushButton("Update")
        self.hapus_btn = QPushButton("Hapus")
        self.clear_btn = QPushButton("Bersihkan")
        
        self.tambah_btn.clicked.connect(self.tambah_barang)
        self.update_btn.clicked.connect(self.update_barang)
        self.hapus_btn.clicked.connect(self.hapus_barang)
        self.clear_btn.clicked.connect(self.clear_form)
        
        btn_layout.addWidget(self.tambah_btn)
        btn_layout.addWidget(self.update_btn)
        btn_layout.addWidget(self.hapus_btn)
        btn_layout.addWidget(self.clear_btn)
        
        self.set_button_styles()
        
        self.product_table = QTableWidget()
        self.setup_product_table()
        
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.product_table)
        
        self.tabs.addTab(product_tab, "Manajemen Produk")
    
    def set_button_styles(self):
        """Mengatur style CSS untuk tombol-tombol manajemen produk."""
        self.tambah_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        self.update_btn.setStyleSheet(self.get_button_style("#2196F3"))
        self.hapus_btn.setStyleSheet(self.get_button_style("#f44336"))
        self.clear_btn.setStyleSheet(self.get_button_style("#9E9E9E"))

    def get_button_style(self, color):
        """Helper untuk menghasilkan style tombol."""
        return f"""
            QPushButton {{ 
                background-color: {color}; 
                color: white; 
                padding: 24px 32px; 
                border-radius: 12px; 
                font-weight: bold; 
                font-size: 18px;
            }} 
            QPushButton:hover {{ 
                background-color: {self.darken_color(color)}; 
            }}
        """

    def darken_color(self, hex_color, factor=0.85):
        """Mendemkan warna hex untuk efek hover."""
        color = QColor(hex_color)
        return color.darker(100 + int(100 * (1 - factor)))
    
    def setup_product_table(self):
        """Mengatur struktur dan tampilan tabel produk."""
        self.product_table.setColumnCount(5)
        self.product_table.setHorizontalHeaderLabels(["Kode", "Nama", "Stok", "Harga Beli", "Harga Jual"])
        self.product_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) 
        self.product_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.product_table.cellClicked.connect(self.on_product_selected)
        
        self.product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents) 
        self.product_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents) 
        
        self.product_table.setColumnWidth(0, 150) 
        self.product_table.setColumnWidth(1, 250) 
        self.product_table.setColumnWidth(2, 100) 
        self.product_table.setColumnWidth(3, 180) 
        self.product_table.setColumnWidth(4, 180) 
    
    # --- Tab Transaksi ---
    def create_transaction_tab(self):
        """Membuat tab untuk melihat riwayat transaksi."""
        transaction_tab = QWidget()
        layout = QVBoxLayout(transaction_tab)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(15) 
        
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        filter_btn = QPushButton("Filter")
        filter_btn.clicked.connect(self.filter_transactions)
        
        filter_btn.setStyleSheet(self.get_button_style("#FF9800"))
        
        filter_layout.addWidget(QLabel("Dari:"))
        filter_layout.addWidget(self.date_from)
        filter_layout.addWidget(QLabel("Sampai:"))
        filter_layout.addWidget(self.date_to)
        filter_layout.addWidget(filter_btn)
        filter_layout.addStretch() 

        self.transaction_table = QTableWidget()
        self.setup_transaction_table()
        
        layout.addLayout(filter_layout)
        layout.addWidget(self.transaction_table)
        
        self.tabs.addTab(transaction_tab, "Transaksi")
    
    def setup_transaction_table(self):
        """Mengatur struktur dan tampilan tabel transaksi."""
        self.transaction_table.setColumnCount(12) 
        self.transaction_table.setHorizontalHeaderLabels(
            ["ID Penjualan", "ID Pesanan", "Tanggal", "Pelanggan", "Kode", "Nama", 
             "Qty", "Harga", "Total", "Metode", "Rating", "Komentar"]
        )
        self.transaction_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.transaction_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents) 
        self.transaction_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents) 
        self.transaction_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents) 
        self.transaction_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents) 
        self.transaction_table.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeToContents) 
        
        self.transaction_table.setColumnWidth(0, 80)  
        self.transaction_table.setColumnWidth(1, 150) 
        self.transaction_table.setColumnWidth(2, 150) 
        self.transaction_table.setColumnWidth(3, 180) 
        self.transaction_table.setColumnWidth(4, 100) 
        self.transaction_table.setColumnWidth(5, 200) 
        self.transaction_table.setColumnWidth(6, 60)  
        self.transaction_table.setColumnWidth(7, 120) 
        self.transaction_table.setColumnWidth(8, 140) 
        self.transaction_table.setColumnWidth(9, 100) 
        self.transaction_table.setColumnWidth(10, 70) 
        self.transaction_table.setColumnWidth(11, 280) 
    
    # --- Tab Laporan ---
    def create_report_tab(self):
        """Membuat tab untuk melihat dan mengekspor laporan keuangan."""
        report_tab = QWidget()
        layout = QVBoxLayout(report_tab)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        report_filter_layout = QHBoxLayout()
        report_filter_layout.setSpacing(15)
        
        self.report_date_from = QDateEdit()
        self.report_date_from.setCalendarPopup(True)
        self.report_date_to = QDateEdit()
        self.report_date_to.setCalendarPopup(True)
        report_filter_btn = QPushButton("Tampilkan Laporan")
        report_filter_btn.clicked.connect(self.update_report)
        
        report_filter_btn.setStyleSheet(self.get_button_style("#9C27B0"))
        
        report_filter_layout.addWidget(QLabel("Periode Laporan:"))
        report_filter_layout.addWidget(self.report_date_from)
        report_filter_layout.addWidget(QLabel("s/d"))
        report_filter_layout.addWidget(self.report_date_to)
        report_filter_layout.addWidget(report_filter_btn)
        report_filter_layout.addStretch()
        
        export_layout = QHBoxLayout()
        export_layout.setSpacing(15)
        self.setup_export_buttons(export_layout)
        
        self.figure = Figure(figsize=(8, 6))
        self.canvas = FigureCanvas(self.figure)
        
        self.profit_table = QTableWidget()
        self.setup_profit_table()
        
        layout.addLayout(report_filter_layout)
        layout.addWidget(self.canvas)
        layout.addWidget(self.profit_table)
        layout.addLayout(export_layout)
        
        self.tabs.addTab(report_tab, "Laporan")
    
    def setup_export_buttons(self, layout):
        """Mengatur tombol ekspor laporan (Excel, PDF)."""
        self.export_excel_btn = QPushButton("📈 Export Excel")
        self.export_pdf_btn = QPushButton("📄 Export PDF")
        
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        
        self.export_excel_btn.setStyleSheet(self.get_button_style("#4CAF50"))
        self.export_pdf_btn.setStyleSheet(self.get_button_style("#2196F3"))
        
        layout.addWidget(self.export_excel_btn)
        layout.addWidget(self.export_pdf_btn)
        layout.addStretch()
    
    def setup_profit_table(self):
        """Mengatur struktur dan tampilan tabel ringkasan laba rugi."""
        self.profit_table.setColumnCount(4)
        self.profit_table.setHorizontalHeaderLabels(
            ["Total Penjualan", "Total Pembelian", "Laba Kotor", "Laba Bersih"]
        )
        self.profit_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.profit_table.setStyleSheet("""
            QTableWidget { font-size: 15px; font-weight: bold; } 
            QTableWidget::item { padding: 15px; } 
            QHeaderView::section { 
                background-color: #2196F3; 
                color: white; 
                font-weight: bold;
                font-size: 14px; 
                padding: 10px;
                border: none;
            }
        """)
    
    # --- Database Operations ---
    def load_products(self):
        """Memuat data produk dari database ke tabel produk."""
        conn = None 
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT kode, nama, stok, harga_beli, harga_jual FROM barang")
            products = cursor.fetchall()
            
            self.product_table.setRowCount(len(products))
            for row_idx, row_data in enumerate(products):
                for col_idx, col_data in enumerate(row_data):
                    item = QTableWidgetItem(str(col_data))
                    item.setFlags(item.flags() ^ Qt.ItemIsEditable) 
                    
                    if col_idx == 0: # Kode
                        item.setTextAlignment(Qt.AlignCenter)
                    elif col_idx == 1: # Nama
                        item.setTextAlignment(Qt.AlignCenter) # Tetap rata kiri
                    elif col_idx == 2: # Stok
                        item.setTextAlignment(Qt.AlignCenter)
                    elif col_idx == 3 or col_idx == 4: # Harga Beli, Harga Jual
                        item.setTextAlignment(Qt.AlignCenter) # Rata kanan untuk angka
                    else: # Default untuk kolom lain jika ada
                        item.setTextAlignment(Qt.AlignCenter)
                    self.product_table.setItem(row_idx, col_idx, item)
                    
            self.set_table_row_colors(self.product_table)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat produk: {str(e)}")
        finally:
            if conn and conn.is_connected():
                conn.close()
    
    def set_table_row_colors(self, table):
        """Mengatur warna baris alternatif untuk tabel."""
        for row in range(table.rowCount()):
            color = QColor(240, 240, 240) if row % 2 == 0 else QColor(255, 255, 255)
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item: 
                    item.setBackground(color)
    
    def filter_transactions(self):
        """Memfilter dan menampilkan transaksi berdasarkan rentang tanggal."""
        conn = None
        try:
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.id, p.id_pesanan, DATE_FORMAT(p.tanggal, '%Y-%m-%d'), p.pelanggan, 
                       p.kode_barang, COALESCE(b.nama, 'Barang dihapus'), p.jumlah, p.harga_satuan, 
                       p.total, p.metode_pembayaran, p.rating, p.komentar
                FROM penjualan p
                LEFT JOIN barang b ON p.kode_barang = b.kode
                WHERE p.tanggal BETWEEN %s AND %s
                ORDER BY p.tanggal DESC
            """, (date_from + " 00:00:00", date_to + " 23:59:59"))
            
            transactions = cursor.fetchall()
            
            self.transaction_table.setColumnCount(12) 
            self.transaction_table.setRowCount(len(transactions))
            
            for row_idx, row_data in enumerate(transactions):
                for col_idx, col_data in enumerate(row_data):
                    item = QTableWidgetItem(str(col_data))
                    item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                    if col_idx in [0, 1, 2, 4, 6, 9, 10]:
                        item.setTextAlignment(Qt.AlignCenter)
                    elif col_idx in [7, 8]: # Harga, Total
                        item.setTextAlignment(Qt.AlignVCenter)
                    else: # Nama Produk, Pelanggan, Komentar
                        item.setTextAlignment(Qt.AlignVCenter)
                    self.transaction_table.setItem(row_idx, col_idx, item)

            self.update_report() 
            self.status_bar.showMessage(f"Menampilkan transaksi dari {date_from} sampai {date_to}")
            
            self.set_table_row_colors(self.transaction_table)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memfilter transaksi: {str(e)}")
        finally:
            if conn and conn.is_connected():
                conn.close()
    
    # --- Fungsi Laporan ---
    def update_report(self):
        """Memperbarui data laporan keuangan dan grafik berdasarkan periode yang dipilih."""
        conn = None
        try:
            self.figure.clear()
            
            start_date = self.report_date_from.date().toString("yyyy-MM-dd")
            end_date = self.report_date_to.date().toString("yyyy-MM-dd")
            
            self.report_data.update({
                'start_date': start_date,
                'end_date': end_date,
                'periode': f"{self.report_date_from.date().toString('dd/MM/yyyy')} - {self.report_date_to.date().toString('dd/MM/yyyy')}"
            })
            
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT SUM(jumlah * harga_satuan) 
                FROM penjualan 
                WHERE tanggal BETWEEN %s AND %s
            """, (start_date + " 00:00:00", end_date + " 23:59:59"))
            total_penjualan = cursor.fetchone()[0] or 0
            
            cursor.execute("""
                SELECT SUM(p.jumlah * b.harga_beli)
                FROM penjualan p
                JOIN barang b ON p.kode_barang = b.kode
                WHERE p.tanggal BETWEEN %s AND %s
            """, (start_date + " 00:00:00", end_date + " 23:59:59"))
            total_pembelian = cursor.fetchone()[0] or 0
            
            laba_kotor = total_penjualan - total_pembelian
            laba_bersih = laba_kotor 
            
            self.report_data.update({
                'total_penjualan': total_penjualan,
                'total_pembelian': total_pembelian,
                'laba_kotor': laba_kotor,
                'laba_bersih': laba_bersih
            })
            
            self.update_profit_table(total_penjualan, total_pembelian, laba_kotor, laba_bersih)
            self.create_profit_chart(total_penjualan, total_pembelian, laba_kotor)
            
            self.status_bar.showMessage(f"Laporan diperbarui untuk periode {self.report_data['periode']}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuat laporan: {str(e)}")
        finally:
            if conn and conn.is_connected():
                conn.close()
    
    def update_profit_table(self, penjualan, pembelian, laba_kotor, laba_bersih):
        """Memperbarui tabel ringkasan laba rugi dengan nilai terbaru."""
        self.profit_table.setRowCount(1)
        values = [penjualan, pembelian, laba_kotor, laba_bersih]
        for col_idx, value in enumerate(values):
            item = QTableWidgetItem(f"Rp{value:,.0f}") 
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.profit_table.setItem(0, col_idx, item)
    
    def create_profit_chart(self, penjualan, pembelian, laba_kotor):
        """Membuat grafik batang (bar chart) untuk visualisasi laba rugi."""
        self.figure.clear()
        
        ax = self.figure.add_subplot(111)
        labels = ['Penjualan', 'Pembelian', 'Laba Kotor']
        values = [penjualan, pembelian, laba_kotor]
        
        colors = ['#4CAF50', '#F44336', '#2196F3']
        bars = ax.bar(labels, values, color=colors)
        ax.set_title(f'Laporan Laba Rugi\nPeriode: {self.report_data["periode"]}')
        ax.set_ylabel('Jumlah (Rp)')
        
        ax.yaxis.set_major_formatter('Rp{x:,.0f}')
        
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                            f'Rp{height:,.0f}',
                            ha='center', va='bottom', fontsize=10)
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    # --- Fungsi Ekspor ---
    def export_to_excel(self):
        """Mengekspor data laporan keuangan ke file Excel."""
        try:
            if not self.report_data['start_date']:
                raise ValueError("Data laporan belum tersedia. Silakan tampilkan laporan terlebih dahulu.")
            
            start_date_fmt = QDate.fromString(self.report_data['start_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            end_date_fmt = QDate.fromString(self.report_data['end_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            periode = f"{start_date_fmt} - {end_date_fmt}"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Simpan Excel", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan"
            
            ws['A1'] = "Laporan Keuangan"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"Periode: {periode}"
            ws.merge_cells('A2:D2')
            
            headers = ["Keterangan", "Jumlah (Rp)"]
            data = [
                ("Total Penjualan", self.report_data['total_penjualan']),
                ("Total Pembelian", self.report_data['total_pembelian']),
                ("Laba Kotor", self.report_data['laba_kotor']),
                ("Laba Bersih", self.report_data['laba_bersih'])
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=4, column=col, value=header).font = Font(bold=True)
                
            for row, (label, value) in enumerate(data, 5):
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value).number_format = '#,##0'
                
            self.adjust_excel_columns(ws)
                
            wb.save(file_path)
            QMessageBox.information(self, "Sukses", "Laporan berhasil diekspor ke Excel!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal ekspor ke Excel: {str(e)}")
    
    def adjust_excel_columns(self, worksheet):
        """Menyesuaikan lebar kolom di worksheet Excel."""
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                try:
                    if cell.value is not None:
                        cell_value_str = str(cell.value)
                        if len(cell_value_str) > max_length:
                            max_length = len(cell_value_str)
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_pdf(self):
        """Mengekspor laporan keuangan ke file PDF."""
        chart_path = None 
        try:
            if not self.report_data['start_date']:
                raise ValueError("Data laporan belum tersedia. Silakan tampilkan laporan terlebih dahulu.")
            
            start_date_fmt = QDate.fromString(self.report_data['start_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            end_date_fmt = QDate.fromString(self.report_data['end_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            periode = f"{start_date_fmt} - {end_date_fmt}"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Simpan PDF", "", "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            chart_path = self.save_chart_image()
                
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "Laporan Keuangan", 0, 1, 'C')
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 10, f"Periode: {periode}", 0, 1, 'C')
            pdf.ln(10)
            
            pdf.image(chart_path, x=10, w=190) 
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(95, 10, "Keterangan", 1, 0, 'C')
            pdf.cell(95, 10, "Jumlah (Rp)", 1, 1, 'C')
            
            pdf.set_font("Arial", '', 11)
            data = [
                ("Total Penjualan", self.report_data['total_penjualan']),
                ("Total Pembelian", self.report_data['total_pembelian']),
                ("Laba Kotor", self.report_data['laba_kotor']),
                ("Laba Bersih", self.report_data['laba_bersih'])
            ]
            
            for label, value in data:
                pdf.cell(95, 10, label, 1, 0, 'L')
                pdf.cell(95, 10, f"{value:,.0f}", 1, 1, 'R')
                
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 8)
            pdf.cell(0, 10, f"Dicetak pada: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", 0, 0, 'R')
            
            pdf.output(file_path)
            QMessageBox.information(self, "Sukses", "Laporan berhasil diekspor ke PDF!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal ekspor ke PDF: {str(e)}")
        finally:
            if chart_path and os.path.exists(chart_path):
                os.unlink(chart_path)
    
    def save_chart_image(self):
        """Menyimpan grafik Matplotlib sebagai gambar PNG sementara."""
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
            self.figure.savefig(tmpfile.name, dpi=300, bbox_inches='tight')
            return tmpfile.name
    
    # --- Manajemen Produk ---
    def on_product_selected(self, row):
        """Mengisi form input produk saat baris produk di tabel dipilih."""
        self.kode_input.setText(self.product_table.item(row, 0).text())
        self.nama_input.setText(self.product_table.item(row, 1).text())
        self.stok_input.setText(self.product_table.item(row, 2).text())
        self.harga_beli_input.setText(self.product_table.item(row, 3).text())
        self.harga_jual_input.setText(self.product_table.item(row, 4).text())
        
    def clear_form(self):
        """Mengosongkan semua field input form produk."""
        self.kode_input.clear()
        self.nama_input.clear()
        self.stok_input.clear()
        self.harga_beli_input.clear()
        self.harga_jual_input.clear()
        
    def validate_product_input(self):
        """Memvalidasi input dari form produk."""
        fields = {
            "Kode": self.kode_input.text(),
            "Nama": self.nama_input.text(),
            "Stok": self.stok_input.text(),
            "Harga Beli": self.harga_beli_input.text(),
            "Harga Jual": self.harga_jual_input.text()
        }
        
        for field, value in fields.items():
            if not value.strip():
                raise ValueError(f"Field '{field}' harus diisi!")
            
        try:
            stok = int(fields['Stok'])
            harga_beli = float(fields['Harga Beli']) 
            harga_jual = float(fields['Harga Jual']) 
        except ValueError:
            raise ValueError("Stok, Harga Beli, dan Harga Jual harus berupa angka!")
        
        if stok < 0:
            raise ValueError("Stok tidak boleh negatif!")
        if harga_beli < 0:
            raise ValueError("Harga Beli tidak boleh negatif!")
        if harga_jual < 0:
            raise ValueError("Harga Jual tidak boleh negatif!")

        if harga_jual <= harga_beli:
            raise ValueError("Harga jual harus lebih besar dari harga beli!")
        
        return fields
        
    def tambah_barang(self):
        """Menambahkan produk baru ke database."""
        conn = None
        try:
            fields = self.validate_product_input()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("SELECT COUNT(*) FROM barang WHERE kode = %s", (fields['Kode'],))
            if cursor.fetchone()[0] > 0:
                raise ValueError("Kode barang sudah ada!")
            
            cursor.execute("""
                INSERT INTO barang (kode, nama, stok, harga_beli, harga_jual)
                VALUES (%s, %s, %s, %s, %s)
            """, (fields['Kode'], fields['Nama'], int(fields['Stok']), 
                  float(fields['Harga Beli']), float(fields['Harga Jual'])))
            conn.commit()
            
            self.load_products()
            self.clear_form()
            QMessageBox.information(self, "Sukses", "Barang berhasil ditambahkan!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
        finally:
            if conn and conn.is_connected():
                conn.close()
            
    def update_barang(self):
        """Memperbarui data produk yang ada di database."""
        conn = None
        try:
            selected_row = self.product_table.currentRow()
            if selected_row == -1:
                raise ValueError("Pilih barang yang akan diupdate!")
                
            fields = self.validate_product_input()
            current_kode = self.product_table.item(selected_row, 0).text()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            if fields['Kode'] != current_kode:
                cursor.execute("SELECT COUNT(*) FROM barang WHERE kode = %s", (fields['Kode'],))
                if cursor.fetchone()[0] > 0:
                    raise ValueError("Kode barang baru sudah ada untuk barang lain!")

            cursor.execute("""
                UPDATE barang 
                SET kode = %s, nama = %s, stok = %s, 
                    harga_beli = %s, harga_jual = %s 
                WHERE kode = %s
            """, (fields['Kode'], fields['Nama'], int(fields['Stok']),
                  float(fields['Harga Beli']), float(fields['Harga Jual']), current_kode))
            conn.commit()
            
            self.load_products()
            self.clear_form()
            QMessageBox.information(self, "Sukses", "Barang berhasil diupdate!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
        finally:
            if conn and conn.is_connected():
                conn.close()
            
    def hapus_barang(self):
        """Menghapus produk dari database."""
        conn = None
        try:
            selected_row = self.product_table.currentRow()
            if selected_row == -1:
                raise ValueError("Pilih barang yang akan dihapus!")
                
            kode = self.product_table.item(selected_row, 0).text()
            nama = self.product_table.item(selected_row, 1).text()
            
            confirm = QMessageBox.question(
                self, "Konfirmasi", 
                f"Apakah Anda yakin ingin menghapus {nama} ({kode})?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if confirm == QMessageBox.Yes:
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM barang WHERE kode = %s", (kode,))
                conn.commit()
                
                self.load_products()
                self.clear_form()
                QMessageBox.information(self, "Sukses", "Barang berhasil dihapus!")
                
        except mysql.connector.IntegrityError as e:
            QMessageBox.critical(self, "Error", 
                                 "Tidak dapat menghapus barang karena ada transaksi terkait di riwayat penjualan.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
        finally:
            if conn and conn.is_connected():
                conn.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    font = QFont()
    font.setFamily("Segoe UI") 
    font.setPointSize(11) 
    app.setFont(font)
    
    window = AdminApp()
    window.show()
    sys.exit(app.exec_())